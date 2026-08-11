from datetime import datetime
from pathlib import Path

import openpyxl

from automation.discovery import analyze_excel


def create_workbook(path: Path) -> None:
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Operations"
    sheet.append(["Created", "Amount", "Status"])
    sheet.append([datetime(2026, 1, 1), 10.5, "Open"])
    sheet.append([datetime(2026, 1, 2), 20.0, "Closed"])
    sheet["E1"] = "=SUM(B2:B3)"
    sheet.merge_cells("G1:H1")
    workbook.create_sheet("Hidden").sheet_state = "hidden"
    workbook.save(path)


def test_excel_analysis_with_permission_returns_evidence_not_values(tmp_path: Path) -> None:
    path = tmp_path / "reference.xlsx"
    create_workbook(path)

    manifest = analyze_excel(path, permit_data_extraction=True)

    assert manifest.format == "xlsx"
    assert manifest.sections[0].name == "Operations"
    assert manifest.sections[0].formula_cells == 1
    assert manifest.sections[0].merged_ranges == 1
    assert manifest.sections[1].hidden is True
    assert [field.name for field in manifest.sections[0].fields] == [
        "Created",
        "Amount",
        "Status",
    ]
    serialized = manifest.model_dump_json()
    assert "10.5" not in serialized
    assert "Closed" not in serialized


def test_excel_analysis_without_permission_withholds_labels(tmp_path: Path) -> None:
    path = tmp_path / "reference.xlsx"
    create_workbook(path)

    manifest = analyze_excel(path, permit_data_extraction=False)

    assert manifest.sections[0].name == "Sheet 1"
    assert manifest.sections[0].fields == []
    assert "Operations" not in manifest.model_dump_json()
    assert manifest.assumptions
