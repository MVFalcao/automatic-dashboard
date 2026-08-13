from pathlib import Path

import pytest

from automation.reports import ArtifactStore, ReportDocument, ReportRequest
from automation.reports.renderers import excel_metric_values, render_excel, render_html, render_pdf
from openpyxl import load_workbook
from io import BytesIO
from automation.specification.models import OutputKind
from test_dashboard_spec import valid_spec


def document() -> ReportDocument:
    spec = valid_spec()
    payload = spec.model_dump(mode="json")
    payload["sections"].append({
        "id": "details", "title": "Details", "kind": "table",
        "field_ids": ["category", "amount"], "depends_on": ["summary"], "order": 1,
    })
    return ReportDocument(
        specification=type(spec).model_validate(payload),
        records=[{"category": "A", "amount": 10}, {"category": "B", "amount": 20}],
        metrics={"total": 30},
        synthetic=True,
    )


def test_excel_and_html_have_metric_terminology_and_section_parity() -> None:
    report = document()
    workbook = render_excel(report)
    html = render_html(report)

    assert excel_metric_values(workbook) == {"Total": 30}
    assert 'data-metric="total">30<' in html
    assert 'data-section="summary">Summary<' in html
    assert "Synthetic data" in html


def test_excel_escapes_formula_like_text() -> None:
    report = document().model_copy(update={"records": [{"category": "=HYPERLINK(\"https://attacker\")", "amount": 10}]})
    workbook = load_workbook(BytesIO(render_excel(report)), data_only=False)
    try:
        assert workbook["Data"]["A2"].value == "'=HYPERLINK(\"https://attacker\")"
        assert workbook["Data"]["A2"].data_type == "s"
    finally:
        workbook.close()


def test_excel_escapes_formula_like_title_headings_labels_and_explanations() -> None:
    spec = document().specification
    payload = spec.model_dump(mode="json")
    payload["title"] = "=TITLE"
    payload["fields"][0]["label"] = "+FIELD"
    payload["metrics"][0]["label"] = "-METRIC"
    payload["metrics"][0]["explanation"] = "@EXPLANATION"
    report = document().model_copy(update={"specification": type(spec).model_validate(payload)})
    workbook = load_workbook(BytesIO(render_excel(report)), data_only=False)
    try:
        assert workbook["Summary"]["A1"].value == "'=TITLE"
        assert workbook["Summary"]["A3"].value == "'-METRIC"
        assert workbook["Summary"]["C3"].value == "'@EXPLANATION"
        assert workbook["Data"]["A1"].value == "'+FIELD"
    finally:
        workbook.close()


def test_pdf_is_printed_from_shared_html() -> None:
    content = render_pdf(document())

    assert content.startswith(b"%PDF")
    assert len(content) > 1_000


def test_confidential_artifact_requires_lifecycle_approval_and_is_deleted_after_transfer(tmp_path: Path) -> None:
    store = ArtifactStore(tmp_path)
    spec = valid_spec()
    payload = spec.model_dump(mode="json")
    payload["fields"][0]["confidential"] = True
    payload["privacy"]["confidential_fields"] = ["category"]
    payload["privacy"]["allow_persistence"] = False
    confidential_document = document().model_copy(update={"specification": type(spec).model_validate(payload)})
    request = ReportRequest(document=confidential_document, outputs=[OutputKind.EXCEL], confidential=True)
    with pytest.raises(ValueError, match="lifecycle approval"):
        store.generate(request)

    approved = request.model_copy(update={"confidential_lifecycle_approved": True})
    artifact = store.generate(approved)[0]
    source_path = tmp_path / f"{artifact.id}.xlsx"
    assert source_path.exists()

    content, deleted_path = store.consume(artifact.id)
    assert content.startswith(b"PK")
    assert deleted_path == source_path
    assert not source_path.exists()
    with pytest.raises(KeyError):
        store.consume(artifact.id)


def test_confidentiality_is_derived_from_approved_specification() -> None:
    with pytest.raises(ValueError, match="must match"):
        ReportRequest(document=document(), outputs=[OutputKind.EXCEL], confidential=True)
