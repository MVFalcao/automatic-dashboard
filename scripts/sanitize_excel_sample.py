"""Create a synthetic, publishable workbook from the private reference template."""

from __future__ import annotations

import argparse
import re
import zipfile
from datetime import datetime, timedelta
from pathlib import Path

import openpyxl
from openpyxl.workbook.properties import CalcProperties


def replace_text(value: str, replacements: dict[str, str]) -> str:
    result = value
    for source, replacement in replacements.items():
        result = re.sub(re.escape(source), replacement, result, flags=re.IGNORECASE)
    return result


def discover_team_replacements(workbook: openpyxl.Workbook) -> dict[str, str]:
    """Discover person-specific team labels from the reference legend."""
    if "legenda" not in workbook.sheetnames:
        return {}
    legend = workbook["legenda"]
    team_column = None
    for cell in legend[1]:
        if isinstance(cell.value, str) and "equipe" in cell.value.casefold():
            team_column = cell.column
            break
    if team_column is None:
        return {}

    names = []
    for row in range(2, legend.max_row + 1):
        value = legend.cell(row, team_column).value
        if not value:
            break
        if isinstance(value, str):
            names.append(value)
    return {name: f"Team {chr(65 + index)}" for index, name in enumerate(names)}


def clear_private_content(workbook: openpyxl.Workbook) -> None:
    """Remove source records while retaining template formulas and formatting."""
    database = workbook["Database"]
    for row in database.iter_rows(min_row=2, max_row=database.max_row, min_col=1, max_col=8):
        for cell in row:
            cell.value = None

    # These columns can contain manually entered assignments, comments, or outcomes.
    candidates = workbook["Candidatos"]
    for row in candidates.iter_rows(min_row=2, max_row=candidates.max_row):
        for column in (10, 11, 12, 13, 14, 15):
            row[column - 1].value = None


def add_synthetic_records(workbook: openpyxl.Workbook, count: int = 36) -> None:
    database = workbook["Database"]
    regions = ("SP", "RJ", "MG", "BA", "PR", "PE")
    cohorts = ("Online Morning", "Online Evening", "In Person")
    incomes = (
        "Up to 1 minimum wage",
        "1 to 2 minimum wages",
        "2 to 3 minimum wages",
        "Above 3 minimum wages",
        "Not provided",
    )
    start = datetime(2026, 1, 5)

    for index in range(1, count + 1):
        row = index + 1
        database.cell(row, 1, start + timedelta(days=index * 2))
        database.cell(row, 2, f"Person {index:03d}")
        database.cell(row, 3, 16 + (index * 3) % 39)
        database.cell(row, 4, f"person{index:03d}@example.invalid")
        database.cell(row, 5, f"+55 00 90000-{index:04d}")
        database.cell(row, 6, regions[(index - 1) % len(regions)])
        database.cell(row, 7, cohorts[(index - 1) % len(cohorts)])
        database.cell(row, 8, incomes[(index - 1) % len(incomes)])


def sanitize_workbook(
    source: Path,
    destination: Path,
    extra_replacements: dict[str, str] | None = None,
) -> list[str]:
    workbook = openpyxl.load_workbook(source, keep_links=False)

    clear_private_content(workbook)

    team_replacements = discover_team_replacements(workbook)
    replacements = {
        **team_replacements,
        **(extra_replacements or {}),
    }

    for old_name, new_name in team_replacements.items():
        if old_name in workbook.sheetnames:
            workbook[old_name].title = new_name

    for sheet in workbook.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                if isinstance(cell.value, str):
                    cell.value = replace_text(cell.value, replacements)

    add_synthetic_records(workbook)

    workbook.properties.creator = "Example Organization"
    workbook.properties.lastModifiedBy = "Example Organization"
    workbook.properties.title = "Synthetic Dashboard Example"
    workbook.properties.subject = "Public synthetic dashboard template"
    workbook.properties.description = (
        "Synthetic example generated for documentation and testing."
    )
    workbook.properties.keywords = "synthetic,dashboard,example"
    workbook.properties.category = "Public example"
    if workbook.calculation is None:
        workbook.calculation = CalcProperties()
    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True

    destination.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(destination)
    return list(replacements)


def scan_workbook(path: Path, forbidden_terms: list[str] | None = None) -> list[str]:
    findings: list[str] = []
    with zipfile.ZipFile(path) as archive:
        for member in archive.namelist():
            if member.lower().endswith((".xml", ".rels")):
                text = archive.read(member).decode("utf-8", errors="ignore")
                for term in forbidden_terms or []:
                    if re.search(re.escape(term), text, flags=re.IGNORECASE):
                        findings.append(f"{member}: forbidden term {term!r}")
                if "externalLink" in member or "TargetMode=\"External\"" in text:
                    findings.append(f"{member}: external relationship")
    return findings


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("source", type=Path)
    parser.add_argument("destination", type=Path)
    parser.add_argument(
        "--replace",
        action="append",
        default=[],
        metavar="SOURCE=REPLACEMENT",
        help="replace a private label; may be supplied more than once",
    )
    parser.add_argument(
        "--forbidden-term",
        action="append",
        default=[],
        help="fail if this term remains inside the generated XLSX",
    )
    args = parser.parse_args()

    replacements = {}
    for item in args.replace:
        if "=" not in item:
            parser.error("--replace must use SOURCE=REPLACEMENT")
        source, replacement = item.split("=", 1)
        replacements[source] = replacement

    discovered_terms = sanitize_workbook(args.source, args.destination, replacements)
    findings = scan_workbook(
        args.destination,
        forbidden_terms=list(dict.fromkeys(discovered_terms + args.forbidden_term)),
    )
    if findings:
        args.destination.unlink(missing_ok=True)
        raise SystemExit("Privacy scan failed:\n" + "\n".join(findings))
    print(f"Created sanitized workbook: {args.destination}")


if __name__ == "__main__":
    main()
