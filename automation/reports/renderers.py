"""Excel and shared HTML/PDF renderers driven only by ReportDocument."""

from __future__ import annotations

import html
from datetime import date, datetime
from io import BytesIO
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from playwright.sync_api import sync_playwright

from automation.reports.models import ReportDocument
from automation.reports.localization import excel_number_format, format_value
from automation.specification.models import FieldKind, SectionKind


SYNTHETIC_NOTICE = {
    "en": "Synthetic data — all values are invented for design review.",
    "pt": "Dados sintéticos — todos os valores são inventados para revisão do design.",
}


def _excel_safe(value: object) -> object:
    if isinstance(value, str) and value[:1] in {"=", "+", "-", "@"}:
        return "'" + value
    return value


def _language(document: ReportDocument) -> str:
    return document.specification.localization.language.split("-")[0]


def _ordered_sections(document: ReportDocument):
    return sorted(document.specification.sections, key=lambda item: item.order)


def _visible_metrics(document: ReportDocument):
    identifiers = {identifier for section in _ordered_sections(document) if section.kind in {SectionKind.SUMMARY, SectionKind.METRICS} for identifier in section.metric_ids}
    return [metric for metric in document.specification.metrics if metric.id in identifiers]


def _visible_fields(document: ReportDocument):
    identifiers = {identifier for section in _ordered_sections(document) if section.kind is SectionKind.TABLE for identifier in section.field_ids}
    return [field for field in document.specification.fields if field.id in identifiers]


def _typed_excel(value: object, kind: FieldKind) -> object:
    if value is None:
        return None
    if kind is FieldKind.NUMBER and isinstance(value, (int, float)) and not isinstance(value, bool):
        return value
    if kind in {FieldKind.DATE, FieldKind.DATETIME} and isinstance(value, str):
        try:
            parsed = datetime.fromisoformat(value.replace("Z", "+00:00"))
            return parsed.date() if kind is FieldKind.DATE else parsed.replace(tzinfo=None)
        except ValueError:
            return _excel_safe(value)
    if isinstance(value, (date, datetime)):
        return value
    return _excel_safe(value)


def render_excel(document: ReportDocument) -> bytes:
    language = _language(document)
    workbook = Workbook()
    accent = document.specification.style.palette[0].lstrip("#")
    summary = workbook.active
    summary.title = "Resumo" if language == "pt" else "Summary"
    summary.append([_excel_safe(document.specification.title)])
    summary["A1"].font = Font(size=18, bold=True, color="FFFFFF")
    summary["A1"].fill = PatternFill("solid", fgColor=accent)
    if document.synthetic:
        summary.append([_excel_safe(SYNTHETIC_NOTICE[language])])
    for metric in _visible_metrics(document):
        summary.append([_excel_safe(metric.label), document.metrics.get(metric.id), _excel_safe(metric.explanation)])
        if isinstance(document.metrics.get(metric.id), (int, float)):
            summary.cell(summary.max_row, 2).number_format = excel_number_format(language=language)
    summary.column_dimensions["A"].width = 34
    summary.column_dimensions["B"].width = 18
    summary.column_dimensions["C"].width = 64

    fields = _visible_fields(document)
    if fields:
        details = workbook.create_sheet("Dados" if language == "pt" else "Data")
        details.append([_excel_safe(field.label) for field in fields])
        for cell in details[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor=accent)
        for record in document.records:
            details.append([_typed_excel(record.get(field.id), field.kind) for field in fields])
            for column, field in enumerate(fields, start=1):
                cell = details.cell(details.max_row, column)
                if field.kind is FieldKind.NUMBER and isinstance(cell.value, (int, float)):
                    cell.number_format = excel_number_format(language=language)
                elif field.kind is FieldKind.DATE and isinstance(cell.value, (date, datetime)):
                    cell.number_format = excel_number_format(language=language, date_only=True)
                elif field.kind is FieldKind.DATETIME and isinstance(cell.value, datetime):
                    cell.number_format = excel_number_format(language=language, datetime_value=True)
        details.freeze_panes = "A2"
        details.auto_filter.ref = details.dimensions
    sections = workbook.create_sheet("Seções" if language == "pt" else "Sections")
    sections.append(["Seção" if language == "pt" else "Section"])
    for section in _ordered_sections(document):
        sections.append([_excel_safe(section.title)])
    for insight in document.insights:
        for value in insight.values():
            sections.append([_excel_safe(value)])
    destination = BytesIO()
    workbook.save(destination)
    workbook.close()
    return destination.getvalue()


def render_html(document: ReportDocument) -> str:
    language = _language(document)
    notice = f'<p class="notice">{html.escape(SYNTHETIC_NOTICE[language])}</p>' if document.synthetic else ""
    metrics = "".join(
        f'<article><span>{html.escape(metric.label)}</span><strong data-metric="{html.escape(metric.id)}">{html.escape(format_value(document.metrics.get(metric.id), language=language))}</strong><small>{html.escape(metric.explanation)}</small></article>'
        for metric in _visible_metrics(document)
    )
    fields = _visible_fields(document)
    headers = "".join(f"<th>{html.escape(field.label)}</th>" for field in fields)
    rows = "".join("<tr>" + "".join(f"<td>{html.escape(format_value(record.get(field.id), language=language))}</td>" for field in fields) + "</tr>" for record in document.records)
    table = f"<table><thead><tr>{headers}</tr></thead><tbody>{rows}</tbody></table>" if fields else ""
    sections = "".join(f'<h2 data-section="{html.escape(section.id)}">{html.escape(section.title)}</h2>' for section in _ordered_sections(document))
    accent = html.escape(document.specification.style.palette[0])
    font = html.escape(document.specification.style.font_family)
    border = html.escape(document.specification.style.border_color)
    spacing = document.specification.style.spacing
    return f'''<!doctype html><html lang="{language}"><head><meta charset="utf-8"><style>
@page{{size:{html.escape(document.specification.outputs.pdf_page_size)};margin:16mm}}body{{font-family:{font},sans-serif;color:#17221b}}h1{{font-size:26px}}.notice{{padding:{spacing}px;background:#fff4cf}}.metrics{{display:grid;grid-template-columns:repeat(3,1fr);gap:{spacing}px}}article{{padding:14px;border:1px solid {border}}}article span,article small{{display:block;color:#667168}}article strong{{display:block;font-size:22px;margin:6px 0}}table{{width:100%;border-collapse:collapse;font-size:9px}}th,td{{padding:6px;border-bottom:1px solid {border};text-align:left}}th{{background:{accent};color:white}}
</style></head><body><h1>{html.escape(document.specification.title)}</h1>{notice}<div class="metrics">{metrics}</div>{sections}{table}</body></html>'''


def render_pdf(document: ReportDocument) -> bytes:
    source = render_html(document)
    with sync_playwright() as playwright:
        browser = playwright.chromium.launch(headless=True)
        try:
            page = browser.new_page()
            page.set_content(source, wait_until="load")
            return page.pdf(print_background=True, prefer_css_page_size=True)
        finally:
            browser.close()


def excel_metric_values(content: bytes) -> dict[str, int | float | None]:
    """Test/diagnostic helper that reads renderer values without desktop Excel."""
    workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
    try:
        sheet = workbook.worksheets[0]
        start = 3 if sheet["A2"].value in SYNTHETIC_NOTICE.values() else 2
        return {str(sheet.cell(row, 1).value): sheet.cell(row, 2).value for row in range(start, sheet.max_row + 1)}
    finally:
        workbook.close()
