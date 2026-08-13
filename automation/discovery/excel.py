"""Permission-aware structural analysis for Excel dashboard samples."""

from __future__ import annotations

from collections import Counter
from datetime import date, datetime
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.cell.cell import Cell

from automation.discovery.models import (
    Confidence,
    FieldEvidence,
    FieldType,
    SampleManifest,
    SectionEvidence,
    StyleEvidence,
)


def _value_type(value: Any) -> FieldType:
    if isinstance(value, bool):
        return FieldType.BOOLEAN
    if isinstance(value, datetime):
        return FieldType.DATETIME
    if isinstance(value, date):
        return FieldType.DATE
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return FieldType.NUMBER
    if isinstance(value, str):
        return FieldType.TEXT
    return FieldType.UNKNOWN


def _header_row(sheet: openpyxl.worksheet.worksheet.Worksheet) -> int | None:
    candidates: list[tuple[int, int]] = []
    for row_number in range(1, min(sheet.max_row, 20) + 1):
        text_cells = sum(
            1
            for cell in sheet[row_number]
            if isinstance(cell.value, str) and cell.value and not cell.value.startswith("=")
        )
        candidates.append((text_cells, row_number))
    count, row_number = max(candidates, default=(0, 0))
    return row_number if count >= 2 else None


def _infer_fields(
    sheet: openpyxl.worksheet.worksheet.Worksheet,
    header_row: int,
) -> list[FieldEvidence]:
    fields: list[FieldEvidence] = []
    for column in range(1, sheet.max_column + 1):
        header = sheet.cell(header_row, column).value
        if not isinstance(header, str) or not header.strip() or header.startswith("="):
            continue
        observed: list[FieldType] = []
        for row in range(header_row + 1, min(sheet.max_row, header_row + 50) + 1):
            value = sheet.cell(row, column).value
            if value is None or (isinstance(value, str) and value.startswith("=")):
                continue
            observed.append(_value_type(value))
        counts = Counter(observed)
        inferred = counts.most_common(1)[0][0] if counts else FieldType.UNKNOWN
        confidence = Confidence.HIGH if observed and counts[inferred] == len(observed) else Confidence.MEDIUM
        if not observed:
            confidence = Confidence.LOW
        fields.append(
            FieldEvidence(
                name=header.strip(),
                source_location=f"{sheet.title}!{sheet.cell(header_row, column).coordinate}",
                inferred_type=inferred,
                confidence=confidence,
                non_empty_samples=len(observed),
            )
        )
    return fields


def _has_formula(cell: Cell) -> bool:
    return cell.data_type == "f" or (
        isinstance(cell.value, str) and cell.value.startswith("=")
    )


def analyze_excel(path: Path, *, permit_data_extraction: bool) -> SampleManifest:
    workbook = openpyxl.load_workbook(path, data_only=False, keep_links=False)
    sections: list[SectionEvidence] = []
    warnings: list[str] = []
    colors: Counter[str] = Counter()
    fonts: Counter[str] = Counter()
    sizes: Counter[float] = Counter()
    borders: Counter[str] = Counter()
    row_heights: Counter[float] = Counter()
    column_widths: Counter[float] = Counter()
    chart_types: Counter[str] = Counter()
    try:
        for index, sheet in enumerate(workbook.worksheets, start=1):
            non_empty = 0
            formulas = 0
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.value is not None:
                        non_empty += 1
                    if _has_formula(cell):
                        formulas += 1
                    if cell.has_style:
                        for color in (cell.fill.fgColor, cell.font.color):
                            if color is not None and color.type == "rgb" and color.rgb:
                                colors[str(color.rgb)[-6:].upper()] += 1
                        if cell.font.name:
                            fonts[cell.font.name] += 1
                        if cell.font.sz:
                            sizes[float(cell.font.sz)] += 1
                        for side in (cell.border.left, cell.border.right, cell.border.top, cell.border.bottom):
                            if side.style:
                                borders[str(side.style)] += 1
            row_heights.update(float(value.height) for value in sheet.row_dimensions.values() if value.height)
            column_widths.update(float(value.width) for value in sheet.column_dimensions.values() if value.width)
            chart_types.update(type(chart).__name__ for chart in sheet._charts)

            fields: list[FieldEvidence] = []
            display_name = f"Sheet {index}"
            source_location = f"worksheet[{index}]"
            if permit_data_extraction:
                display_name = sheet.title
                source_location = sheet.title
                header_row = _header_row(sheet)
                if header_row is not None:
                    fields = _infer_fields(sheet, header_row)

            sections.append(
                SectionEvidence(
                    name=display_name,
                    source_location=source_location,
                    rows=sheet.max_row,
                    columns=sheet.max_column,
                    non_empty_cells=non_empty,
                    formula_cells=formulas,
                    merged_ranges=len(sheet.merged_cells.ranges),
                    chart_count=len(sheet._charts),
                    hidden=sheet.sheet_state != "visible",
                    fields=fields,
                )
            )
        if workbook._external_links:
            warnings.append("The workbook contains external links; they were not followed.")
    finally:
        workbook.close()

    assumptions = []
    if not permit_data_extraction:
        assumptions.append(
            "Worksheet names, labels, and field candidates were withheld because data extraction was not permitted."
        )
    return SampleManifest(
        format="xlsx",
        sections=sections,
        extraction_permitted=permit_data_extraction,
        assumptions=assumptions,
        warnings=warnings,
        style=StyleEvidence(
            palette=[f"#{value}" for value, _ in colors.most_common(8)],
            font_families=[value for value, _ in fonts.most_common(4)],
            font_sizes=sorted(value for value, _ in sizes.most_common(6)),
            row_heights=sorted(value for value, _ in row_heights.most_common(6)),
            column_widths=sorted(value for value, _ in column_widths.most_common(6)),
            border_styles=[value for value, _ in borders.most_common(5)],
            chart_types=[value for value, _ in chart_types.most_common(6)],
            organization=[f"worksheet[{index}]" for index, _ in enumerate(sections, start=1)],
            confidence=Confidence.HIGH if sum(colors.values()) >= 5 else Confidence.MEDIUM if colors else Confidence.LOW,
            requires_review=sum(colors.values()) < 5,
        ),
    )
