"""Render consistent synthetic web, Excel, and PDF previews."""

from __future__ import annotations

import base64
import html
import random
from datetime import date, datetime, timedelta, timezone
from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen.canvas import Canvas

from automation.discovery.models import FieldType
from automation.preview.models import (
    PreviewArtifact,
    PreviewLanguage,
    PreviewOutput,
    PreviewPackage,
    PreviewRequest,
)


LABELS = {
    PreviewLanguage.ENGLISH: {
        "title": "Synthetic dashboard preview",
        "notice": "All values in this preview are invented for design review.",
        "records": "Example records",
    },
    PreviewLanguage.PORTUGUESE: {
        "title": "Prévia sintética do dashboard",
        "notice": "Todos os valores desta prévia foram inventados para revisão do design.",
        "records": "Registros de exemplo",
    },
}


def _excel_safe(value: object) -> object:
    return "'" + value if isinstance(value, str) and value[:1] in {"=", "+", "-", "@"} else value


def _synthetic_value(field_id: str, field_type: FieldType, index: int) -> Any:
    seed = f"{field_id}:{index}"
    randomizer = random.Random(seed)
    normalized = field_id.casefold()
    if field_type == FieldType.BOOLEAN:
        return index % 2 == 0
    if field_type == FieldType.NUMBER:
        return round(randomizer.uniform(10, 1_000), 2)
    if field_type == FieldType.DATE:
        return date(2026, 1, 1) + timedelta(days=index * 3)
    if field_type == FieldType.DATETIME:
        return datetime(2026, 1, 1, tzinfo=timezone.utc) + timedelta(days=index, hours=index)
    if "email" in normalized:
        return f"person{index:03d}@example.invalid"
    if "phone" in normalized or "contact" in normalized:
        return f"+00 000 000 {index:04d}"
    if "name" in normalized:
        return f"Example Person {index:03d}"
    return f"Example {index:03d}"


def _records(request: PreviewRequest) -> list[dict[str, Any]]:
    return [
        {
            field.id: _synthetic_value(field.id, field.inferred_type, index)
            for field in request.dashboard_schema.fields
        }
        for index in range(1, request.synthetic_record_count + 1)
    ]


def _display(value: Any) -> str:
    if isinstance(value, (date, datetime)):
        return value.isoformat()
    return str(value)


def _web(request: PreviewRequest, records: list[dict[str, Any]]) -> bytes:
    labels = LABELS[request.language]
    headers = [field.display_name for field in request.dashboard_schema.fields]
    field_ids = [field.id for field in request.dashboard_schema.fields]
    header_html = "".join(f"<th>{html.escape(header)}</th>" for header in headers)
    rows_html = "".join(
        "<tr>" + "".join(
            f"<td>{html.escape(_display(record.get(field_id, '')))}</td>"
            for field_id in field_ids
        ) + "</tr>"
        for record in records
    )
    sections = "".join(
        f"<section><h2>{html.escape(section.display_name)}</h2><p>{html.escape(section.presentation)}</p></section>"
        for section in request.dashboard_schema.sections
    )
    document = f"""<!doctype html>
<html lang="{request.language.value}"><head><meta charset="utf-8"><title>{html.escape(labels['title'])}</title>
<style>body{{font-family:Arial,sans-serif;margin:40px;color:#17221b}}.notice{{padding:12px;background:#eef5ef}}table{{border-collapse:collapse;width:100%}}th,td{{border:1px solid #ccd4ce;padding:8px;text-align:left}}section{{margin:18px 0}}</style></head>
<body><h1>{html.escape(labels['title'])}</h1><p class="notice">{html.escape(labels['notice'])}</p>{sections}<h2>{html.escape(labels['records'])}</h2><table><thead><tr>{header_html}</tr></thead><tbody>{rows_html}</tbody></table></body></html>"""
    return document.encode("utf-8")


def _excel(request: PreviewRequest, records: list[dict[str, Any]]) -> bytes:
    labels = LABELS[request.language]
    workbook = Workbook()
    summary = workbook.active
    summary.title = "Summary" if request.language == PreviewLanguage.ENGLISH else "Resumo"
    summary.append([_excel_safe(labels["title"])])
    summary.append([_excel_safe(labels["notice"])])
    summary["A1"].font = Font(size=16, bold=True, color="FFFFFF")
    summary["A1"].fill = PatternFill("solid", fgColor="23543C")
    summary.column_dimensions["A"].width = 72

    data = workbook.create_sheet("Data" if request.language == PreviewLanguage.ENGLISH else "Dados")
    data.append([_excel_safe(field.display_name) for field in request.dashboard_schema.fields])
    for cell in data[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="23543C")
    for record in records:
        data.append([_excel_safe(record.get(field.id)) for field in request.dashboard_schema.fields])
    data.freeze_panes = "A2"
    data.auto_filter.ref = data.dimensions

    destination = BytesIO()
    workbook.save(destination)
    workbook.close()
    return destination.getvalue()


def _pdf(request: PreviewRequest, records: list[dict[str, Any]]) -> bytes:
    labels = LABELS[request.language]
    destination = BytesIO()
    canvas = Canvas(destination, pagesize=A4)
    width, height = A4
    canvas.setTitle(labels["title"])
    canvas.setFont("Helvetica-Bold", 18)
    canvas.drawString(48, height - 54, labels["title"])
    canvas.setFont("Helvetica", 10)
    canvas.drawString(48, height - 78, labels["notice"])
    y = height - 112
    for section in request.dashboard_schema.sections:
        canvas.setFont("Helvetica-Bold", 12)
        canvas.drawString(48, y, section.display_name[:70])
        y -= 18
        canvas.setFont("Helvetica", 9)
        canvas.drawString(60, y, section.presentation)
        y -= 24
        if y < 80:
            canvas.showPage()
            y = height - 54
    canvas.setFont("Helvetica", 9)
    canvas.drawRightString(width - 48, 40, f"{labels['records']}: {len(records)}")
    canvas.save()
    return destination.getvalue()


def _artifact(output: PreviewOutput, content: bytes) -> PreviewArtifact:
    metadata = {
        PreviewOutput.WEB: ("dashboard-preview.html", "text/html"),
        PreviewOutput.EXCEL: ("dashboard-preview.xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"),
        PreviewOutput.PDF: ("dashboard-preview.pdf", "application/pdf"),
    }
    filename, media_type = metadata[output]
    return PreviewArtifact(
        format=output,
        filename=filename,
        media_type=media_type,
        content_base64=base64.b64encode(content).decode("ascii"),
        size_bytes=len(content),
    )


def generate_preview(request: PreviewRequest) -> PreviewPackage:
    records = _records(request)
    renderers = {
        PreviewOutput.WEB: _web,
        PreviewOutput.EXCEL: _excel,
        PreviewOutput.PDF: _pdf,
    }
    artifacts = [
        _artifact(output, renderers[output](request, records))
        for output in request.outputs
    ]
    return PreviewPackage(
        record_count=len(records),
        artifacts=artifacts,
        warnings=[LABELS[request.language]["notice"]],
    )
