"""Inspect local CSV/Excel data without returning or persisting row values."""

from __future__ import annotations

import csv
import re
from pathlib import Path

import openpyxl

from automation.discovery.models import DraftDashboardSchema
from automation.importing.models import DataSourceSummary, ImportPlan, ProposedMapping


SUPPORTED_SUFFIXES = {".csv", ".xlsx"}
CONFIDENTIAL_PATTERNS = (
    "address",
    "birth",
    "cnpj",
    "contact",
    "cpf",
    "email",
    "endereco",
    "endereço",
    "income",
    "name",
    "nome",
    "phone",
    "renda",
    "salario",
    "salário",
    "telefone",
    "whatsapp",
)


def _normalize(value: str) -> str:
    return re.sub(r"[^a-z0-9]+", "", value.casefold())


def _likely_confidential(columns: list[str]) -> list[str]:
    return [
        column
        for column in columns
        if any(pattern in column.casefold() for pattern in CONFIDENTIAL_PATTERNS)
    ]


def _csv_summary(path: Path) -> DataSourceSummary:
    with path.open("r", encoding="utf-8-sig", newline="") as source:
        reader = csv.reader(source)
        try:
            columns = [value.strip() for value in next(reader)]
        except StopIteration:
            columns = []
        rows = [row for row in reader if any(value.strip() for value in row)]
    issues: list[str] = []
    if not columns:
        issues.append("No header row was found")
    if len(columns) != len(set(columns)):
        issues.append("Column names must be unique")
    if any(len(row) != len(columns) for row in rows):
        issues.append("Some rows have a different number of values than the header")
    return DataSourceSummary(
        filename=path.name,
        format="csv",
        section="rows",
        row_count=len(rows),
        column_count=len(columns),
        columns=columns,
        likely_confidential_columns=_likely_confidential(columns),
        validation_issues=issues,
    )


def _excel_summaries(path: Path) -> list[DataSourceSummary]:
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True, keep_links=False)
    summaries: list[DataSourceSummary] = []
    try:
        for sheet in workbook.worksheets:
            first_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), ())
            columns = [str(value).strip() if value is not None else "" for value in first_row]
            populated_columns = [column for column in columns if column]
            issues: list[str] = []
            if not populated_columns:
                issues.append("No header row was found")
            if len(populated_columns) != len(set(populated_columns)):
                issues.append("Column names must be unique")
            row_count = sum(
                1
                for row in sheet.iter_rows(min_row=2, values_only=True)
                if any(value is not None and value != "" for value in row)
            )
            summaries.append(
                DataSourceSummary(
                    filename=path.name,
                    format="xlsx",
                    section=sheet.title,
                    row_count=row_count,
                    column_count=len(populated_columns),
                    columns=populated_columns,
                    likely_confidential_columns=_likely_confidential(populated_columns),
                    validation_issues=issues,
                )
            )
    finally:
        workbook.close()
    return summaries


def inspect_data_location(path: Path, schema: DraftDashboardSchema) -> ImportPlan:
    if not path.exists():
        raise ValueError("The selected local data location does not exist")
    files = (
        sorted(item for item in path.iterdir() if item.is_file() and item.suffix.casefold() in SUPPORTED_SUFFIXES)
        if path.is_dir()
        else [path]
    )
    if not files or any(item.suffix.casefold() not in SUPPORTED_SUFFIXES for item in files):
        raise ValueError("Dashboard population supports local CSV and XLSX files only")

    summaries: list[DataSourceSummary] = []
    issues: list[str] = []
    for file in files:
        try:
            if file.suffix.casefold() == ".csv":
                summaries.append(_csv_summary(file))
            else:
                summaries.extend(_excel_summaries(file))
        except (UnicodeError, csv.Error, OSError, ValueError) as exc:
            issues.append(f"{file.name} could not be inspected: {type(exc).__name__}")

    issues.extend(
        f"{source.filename}/{source.section}: {issue}"
        for source in summaries
        for issue in source.validation_issues
    )

    target_by_normalized = {_normalize(field.display_name): field.id for field in schema.fields}
    target_by_normalized.update({_normalize(field.id): field.id for field in schema.fields})
    source_columns = list(dict.fromkeys(column for summary in summaries for column in summary.columns))
    mappings = [
        ProposedMapping(
            source_column=column,
            target_field=target_by_normalized.get(_normalize(column)),
            confidence="high" if _normalize(column) in target_by_normalized else "unmapped",
        )
        for column in source_columns
    ]
    return ImportPlan(
        sources=summaries,
        mappings=mappings,
        validation_issues=issues,
        requires_relationship_confirmation=len(summaries) > 1,
    )
